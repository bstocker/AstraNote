"""Modules : grille de notation, dates, colonnes, groupes, saisie des étoiles.

C'est le cœur de l'application (fiche §5.3 à §5.7). La vue module affiche un
tableau unique de consultation + saisie ; les étoiles / notes / URL sont
enregistrées immédiatement via des endpoints AJAX avec recalcul du prorata.
"""
from io import BytesIO
import re

from flask import (
    Blueprint, render_template, redirect, url_for, request, flash, abort, jsonify,
    send_file,
)
from flask_login import login_required, current_user

from .models import (
    db, Class, Module, GradeDate, StarColumn, UrlColumn, NoteColumn,
    Student, Enrollment, Group, GroupMember, Star, UrlValue, NoteValue,
    SUBJECT_STUDENT, SUBJECT_GROUP, WORK_MODE_INDIVIDUAL, WORK_MODE_GROUP,
)
from . import grading
from .main import get_class_or_403, purge_subject_data

modules_bp = Blueprint("modules", __name__)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def get_module_or_403(module_id):
    module = db.session.get(Module, module_id) or abort(404)
    get_class_or_403(module.class_id)  # applique le cloisonnement des droits
    return module


def _next_position(items):
    return (max((i.position or 0) for i in items) + 1) if items else 0


def module_subjects(module):
    """Liste ordonnée des unités notées : (subject_type, id, label, enrollment?).

    En mode individuel : les étudiants inscrits. En mode groupe : les groupes.
    """
    if module.is_group_mode:
        groups = sorted(module.groups, key=lambda g: g.name.lower())
        return [
            {"type": SUBJECT_GROUP, "id": g.id, "label": g.name,
             "comment": g.comment, "obj": g, "active": True}
            for g in groups
        ]
    # Actifs d'abord (ordre alphabétique), puis les neutralisés en fin de liste.
    enrollments = sorted(
        module.klass.enrollments,
        key=lambda e: (not e.student.active, e.student.full_name.lower()),
    )
    return [
        {"type": SUBJECT_STUDENT, "id": e.student.id, "label": e.student.full_name,
         "comment": e.general_comment, "obj": e.student, "enrollment": e,
         "active": e.student.active}
        for e in enrollments
    ]


# --------------------------------------------------------------------------- #
# CRUD Module
# --------------------------------------------------------------------------- #
@modules_bp.route("/classes/<int:class_id>/modules/new", methods=["POST"])
@login_required
def create_module(class_id):
    klass = get_class_or_403(class_id)
    name = request.form.get("name", "").strip()
    work_mode = request.form.get("work_mode", WORK_MODE_INDIVIDUAL)
    if work_mode not in (WORK_MODE_INDIVIDUAL, WORK_MODE_GROUP):
        work_mode = WORK_MODE_INDIVIDUAL
    if not name:
        flash("Le nom du module est requis.", "error")
        return redirect(url_for("main.view_class", class_id=class_id))

    module = Module(
        name=name, class_id=klass.id, work_mode=work_mode,
        discord_url=request.form.get("discord_url", "").strip() or None,
    )
    db.session.add(module)
    db.session.commit()
    flash("Module créé.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/modules/<int:module_id>/edit", methods=["POST"])
@login_required
def edit_module(module_id):
    """Édite un module : nom et lien Discord.

    Le mode de travail (individuel/groupe) n'est PAS modifiable après création
    (il conditionne groupes et notes déjà saisis — cf. fiche §2).
    """
    module = get_module_or_403(module_id)
    name = request.form.get("name", "").strip()
    if not name:
        flash("Le nom du module est requis.", "error")
        return redirect(url_for("modules.view_module", module_id=module.id))

    module.name = name
    module.discord_url = request.form.get("discord_url", "").strip() or None
    db.session.commit()
    flash("Module mis à jour.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/modules/<int:module_id>/delete", methods=["POST"])
@login_required
def delete_module(module_id):
    module = get_module_or_403(module_id)
    class_id = module.class_id
    db.session.delete(module)
    db.session.commit()
    flash("Module supprimé.", "success")
    return redirect(url_for("main.view_class", class_id=class_id))


# --------------------------------------------------------------------------- #
# Vue module : la grille
# --------------------------------------------------------------------------- #
@modules_bp.route("/modules/<int:module_id>")
@login_required
def view_module(module_id):
    module = get_module_or_403(module_id)
    subjects = module_subjects(module)
    subject_ids = [s["id"] for s in subjects]
    active_ids = {s["id"] for s in subjects if s.get("active", True)}
    grades = grading.compute_module_grades(module, subject_ids, active_ids)

    # Index des valeurs pour un accès O(1) dans le template.
    subject_type = SUBJECT_GROUP if module.is_group_mode else SUBJECT_STUDENT
    star_map = {}   # (subject_id, star_column_id) -> value
    url_map = {}    # (subject_id, url_column_id) -> url
    note_map = {}   # (subject_id, note_column_id) -> score

    star_col_ids, url_col_ids = [], []
    for gd in module.grade_dates:
        star_col_ids += [c.id for c in gd.star_columns]
        url_col_ids += [c.id for c in gd.url_columns]
    note_col_ids = [c.id for c in module.note_columns]

    if star_col_ids:
        for s in Star.query.filter(
            Star.subject_type == subject_type,
            Star.star_column_id.in_(star_col_ids)).all():
            star_map[(s.subject_id, s.star_column_id)] = s.value
    if url_col_ids:
        for u in UrlValue.query.filter(
            UrlValue.subject_type == subject_type,
            UrlValue.url_column_id.in_(url_col_ids)).all():
            url_map[(u.subject_id, u.url_column_id)] = u.url
    if note_col_ids:
        for n in NoteValue.query.filter(
            NoteValue.subject_type == subject_type,
            NoteValue.note_column_id.in_(note_col_ids)).all():
            note_map[(n.subject_id, n.note_column_id)] = n.score

    # Étudiants actifs encore non affectés à un groupe (mode groupe). Les
    # étudiants neutralisés (partis) n'ont pas besoin d'être affectés.
    unassigned_active = []
    if module.is_group_mode:
        assigned = {
            m.student_id for g in module.groups for m in g.members
        }
        unassigned_active = sorted(
            (e.student for e in module.klass.enrollments
             if e.student.id not in assigned and e.student.active),
            key=lambda s: s.full_name.lower(),
        )

    return render_template(
        "modules/module_detail.html",
        module=module, subjects=subjects, grades=grades,
        star_map=star_map, url_map=url_map, note_map=note_map,
        all_tokens=grading.ALL_TOKENS, special_statuses=grading.SPECIAL_STATUSES,
        unassigned_active=unassigned_active,
    )


# --------------------------------------------------------------------------- #
# Dates & colonnes
# --------------------------------------------------------------------------- #
@modules_bp.route("/modules/<int:module_id>/dates", methods=["POST"])
@login_required
def add_date(module_id):
    module = get_module_or_403(module_id)
    from datetime import datetime
    raw_date = request.form.get("date", "").strip()
    parsed = None
    if raw_date:
        try:
            parsed = datetime.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            parsed = None
    gd = GradeDate(
        module_id=module.id,
        label=request.form.get("label", "").strip() or None,
        date=parsed,
        position=_next_position(module.grade_dates),
    )
    db.session.add(gd)
    db.session.commit()
    flash("Date/séance ajoutée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/dates/<int:date_id>/delete", methods=["POST"])
@login_required
def delete_date(date_id):
    gd = db.session.get(GradeDate, date_id) or abort(404)
    module = get_module_or_403(gd.module_id)
    db.session.delete(gd)
    db.session.commit()
    flash("Date supprimée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/dates/<int:date_id>/star-columns", methods=["POST"])
@login_required
def add_star_column(date_id):
    gd = db.session.get(GradeDate, date_id) or abort(404)
    module = get_module_or_403(gd.module_id)
    db.session.add(StarColumn(
        grade_date_id=gd.id,
        title=request.form.get("title", "").strip() or "Exercice",
        position=_next_position(gd.star_columns),
    ))
    db.session.commit()
    flash("Colonne d'étoiles ajoutée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/dates/<int:date_id>/url-columns", methods=["POST"])
@login_required
def add_url_column(date_id):
    gd = db.session.get(GradeDate, date_id) or abort(404)
    module = get_module_or_403(gd.module_id)
    db.session.add(UrlColumn(
        grade_date_id=gd.id,
        title=request.form.get("title", "").strip() or "Lien",
        position=_next_position(gd.url_columns),
    ))
    db.session.commit()
    flash("Colonne URL ajoutée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/star-columns/<int:col_id>/delete", methods=["POST"])
@login_required
def delete_star_column(col_id):
    col = db.session.get(StarColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    db.session.delete(col)
    db.session.commit()
    flash("Colonne supprimée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/url-columns/<int:col_id>/delete", methods=["POST"])
@login_required
def delete_url_column(col_id):
    col = db.session.get(UrlColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    db.session.delete(col)
    db.session.commit()
    flash("Colonne supprimée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/modules/<int:module_id>/note-columns", methods=["POST"])
@login_required
def add_note_column(module_id):
    module = get_module_or_403(module_id)
    title = request.form.get("title", "").strip()
    if not title:
        flash("L'intitulé de la colonne de note est requis.", "error")
        return redirect(url_for("modules.view_module", module_id=module.id))
    db.session.add(NoteColumn(
        module_id=module.id, title=title,
        position=_next_position(module.note_columns),
    ))
    db.session.commit()
    flash("Colonne de note ajoutée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/note-columns/<int:col_id>/delete", methods=["POST"])
@login_required
def delete_note_column(col_id):
    col = db.session.get(NoteColumn, col_id) or abort(404)
    module = get_module_or_403(col.module_id)
    db.session.delete(col)
    db.session.commit()
    flash("Colonne de note supprimée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


# --------------------------------------------------------------------------- #
# Renommage, réorganisation et édition (dates & colonnes) — fiche §5.3
# --------------------------------------------------------------------------- #
def _reorder(item, siblings, direction):
    """Normalise les positions puis échange `item` avec son voisin (up/down)."""
    ordered = sorted(siblings, key=lambda x: (x.position or 0, x.id))
    for i, s in enumerate(ordered):
        s.position = i
    idx = ordered.index(item)
    j = idx + (-1 if direction == "up" else 1)
    if 0 <= j < len(ordered):
        ordered[idx].position, ordered[j].position = (
            ordered[j].position, ordered[idx].position,
        )


@modules_bp.route("/star-columns/<int:col_id>/rename", methods=["POST"])
@login_required
def rename_star_column(col_id):
    col = db.session.get(StarColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    title = request.form.get("title", "").strip()
    if title:
        col.title = title
        db.session.commit()
        flash("Colonne renommée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/star-columns/<int:col_id>/move", methods=["POST"])
@login_required
def move_star_column(col_id):
    col = db.session.get(StarColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    _reorder(col, col.grade_date.star_columns, request.form.get("dir", "up"))
    db.session.commit()
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/url-columns/<int:col_id>/rename", methods=["POST"])
@login_required
def rename_url_column(col_id):
    col = db.session.get(UrlColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    title = request.form.get("title", "").strip()
    if title:
        col.title = title
        db.session.commit()
        flash("Colonne renommée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/url-columns/<int:col_id>/move", methods=["POST"])
@login_required
def move_url_column(col_id):
    col = db.session.get(UrlColumn, col_id) or abort(404)
    module = get_module_or_403(col.grade_date.module_id)
    _reorder(col, col.grade_date.url_columns, request.form.get("dir", "up"))
    db.session.commit()
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/note-columns/<int:col_id>/rename", methods=["POST"])
@login_required
def rename_note_column(col_id):
    col = db.session.get(NoteColumn, col_id) or abort(404)
    module = get_module_or_403(col.module_id)
    title = request.form.get("title", "").strip()
    if title:
        col.title = title
        db.session.commit()
        flash("Colonne renommée.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/note-columns/<int:col_id>/move", methods=["POST"])
@login_required
def move_note_column(col_id):
    col = db.session.get(NoteColumn, col_id) or abort(404)
    module = get_module_or_403(col.module_id)
    _reorder(col, module.note_columns, request.form.get("dir", "up"))
    db.session.commit()
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/dates/<int:date_id>/edit", methods=["POST"])
@login_required
def edit_date(date_id):
    gd = db.session.get(GradeDate, date_id) or abort(404)
    module = get_module_or_403(gd.module_id)
    from datetime import datetime
    raw = request.form.get("date", "").strip()
    if raw:
        try:
            gd.date = datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            pass
    gd.label = request.form.get("label", "").strip() or None
    db.session.commit()
    flash("Date mise à jour.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/dates/<int:date_id>/move", methods=["POST"])
@login_required
def move_date(date_id):
    gd = db.session.get(GradeDate, date_id) or abort(404)
    module = get_module_or_403(gd.module_id)
    _reorder(gd, module.grade_dates, request.form.get("dir", "up"))
    db.session.commit()
    return redirect(url_for("modules.view_module", module_id=module.id))


# --------------------------------------------------------------------------- #
# Groupes (mode groupe)
# --------------------------------------------------------------------------- #
@modules_bp.route("/modules/<int:module_id>/groups", methods=["POST"])
@login_required
def add_group(module_id):
    module = get_module_or_403(module_id)
    if not module.is_group_mode:
        abort(400)
    name = request.form.get("name", "").strip()
    if name:
        db.session.add(Group(module_id=module.id, name=name))
        db.session.commit()
        flash("Groupe créé.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/groups/<int:group_id>/members", methods=["POST"])
@login_required
def add_group_member(group_id):
    group = db.session.get(Group, group_id) or abort(404)
    module = get_module_or_403(group.module_id)
    student_id = request.form.get("student_id", type=int)
    if student_id:
        # Un étudiant ne peut appartenir qu'à un seul groupe du module.
        existing_ids = {
            m.student_id for g in module.groups for m in g.members
        }
        if student_id not in existing_ids:
            db.session.add(GroupMember(group_id=group.id, student_id=student_id))
            db.session.commit()
            flash("Étudiant affecté au groupe.", "success")
        else:
            flash("Cet étudiant est déjà dans un groupe.", "error")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/group-members/<int:member_id>/delete", methods=["POST"])
@login_required
def remove_group_member(member_id):
    member = db.session.get(GroupMember, member_id) or abort(404)
    group = member.group
    module = get_module_or_403(group.module_id)
    db.session.delete(member)
    db.session.commit()
    flash("Étudiant retiré du groupe.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


@modules_bp.route("/groups/<int:group_id>/delete", methods=["POST"])
@login_required
def delete_group(group_id):
    group = db.session.get(Group, group_id) or abort(404)
    module = get_module_or_403(group.module_id)
    # Purge les étoiles/notes/liens du groupe (référencés par subject_id).
    purge_subject_data(SUBJECT_GROUP, group.id)
    db.session.delete(group)
    db.session.commit()
    flash("Groupe supprimé.", "success")
    return redirect(url_for("modules.view_module", module_id=module.id))


# --------------------------------------------------------------------------- #
# Saisie AJAX (étoiles, notes, URL, commentaires) + recalcul
# --------------------------------------------------------------------------- #
def _subject_type(module):
    return SUBJECT_GROUP if module.is_group_mode else SUBJECT_STUDENT


def _subject_is_active(module, subject_id):
    """Un étudiant neutralisé n'accepte plus aucune saisie (les groupes, si.)."""
    if module.is_group_mode:
        return True
    student = db.session.get(Student, subject_id)
    return bool(student and student.active)


def _grade_payload(module):
    """Renvoie totaux + notes /20 recalculés pour tout le module."""
    subjects = module_subjects(module)
    subject_ids = [s["id"] for s in subjects]
    active_ids = {s["id"] for s in subjects if s.get("active", True)}
    grades = grading.compute_module_grades(module, subject_ids, active_ids)
    return {
        str(sid): {
            "total": g["total"],
            # "—" pour un sujet neutralisé, "N/A" si personne n'a d'étoile.
            "note": ("—" if not g["active"] else
                     ("N/A" if g["note"] is None else g["note"])),
            "is_reference": g["is_reference"],
        }
        for sid, g in grades.items()
    }


@modules_bp.route("/modules/<int:module_id>/save-star", methods=["POST"])
@login_required
def save_star(module_id):
    module = get_module_or_403(module_id)
    data = request.get_json(silent=True) or {}
    subject_id = data.get("subject_id")
    column_id = data.get("column_id")
    value = str(data.get("value", "0")).strip()

    if value not in grading.ALL_TOKENS:
        return jsonify(error="Valeur invalide"), 400
    col = db.session.get(StarColumn, column_id)
    if not col or col.grade_date.module_id != module.id:
        return jsonify(error="Colonne invalide"), 400
    if not _subject_is_active(module, subject_id):
        return jsonify(error="Étudiant neutralisé : saisie impossible"), 403

    stype = _subject_type(module)
    star = Star.query.filter_by(
        subject_type=stype, subject_id=subject_id, star_column_id=column_id,
    ).first()
    if star:
        star.value = value
    else:
        db.session.add(Star(
            subject_type=stype, subject_id=subject_id,
            star_column_id=column_id, value=value,
        ))
    db.session.commit()

    return jsonify(
        ok=True,
        display=grading.display_token(value),
        color=grading.status_color(value),
        grades=_grade_payload(module),
    )


@modules_bp.route("/modules/<int:module_id>/save-note", methods=["POST"])
@login_required
def save_note(module_id):
    module = get_module_or_403(module_id)
    data = request.get_json(silent=True) or {}
    subject_id = data.get("subject_id")
    column_id = data.get("column_id")
    raw = str(data.get("value", "")).strip().replace(",", ".")

    col = db.session.get(NoteColumn, column_id)
    if not col or col.module_id != module.id:
        return jsonify(error="Colonne invalide"), 400
    if not _subject_is_active(module, subject_id):
        return jsonify(error="Étudiant neutralisé : saisie impossible"), 403

    score = None
    if raw != "":
        try:
            score = float(raw)
        except ValueError:
            return jsonify(error="Note invalide"), 400
        if not (0 <= score <= 20):
            return jsonify(error="La note doit être entre 0 et 20"), 400

    stype = _subject_type(module)
    nv = NoteValue.query.filter_by(
        subject_type=stype, subject_id=subject_id, note_column_id=column_id,
    ).first()
    if nv:
        nv.score = score
    else:
        db.session.add(NoteValue(
            subject_type=stype, subject_id=subject_id,
            note_column_id=column_id, score=score,
        ))
    db.session.commit()
    return jsonify(ok=True, value=("" if score is None else score))


@modules_bp.route("/modules/<int:module_id>/save-url", methods=["POST"])
@login_required
def save_url(module_id):
    module = get_module_or_403(module_id)
    data = request.get_json(silent=True) or {}
    subject_id = data.get("subject_id")
    column_id = data.get("column_id")
    url = str(data.get("value", "")).strip() or None

    col = db.session.get(UrlColumn, column_id)
    if not col or col.grade_date.module_id != module.id:
        return jsonify(error="Colonne invalide"), 400
    if not _subject_is_active(module, subject_id):
        return jsonify(error="Étudiant neutralisé : saisie impossible"), 403

    stype = _subject_type(module)
    uv = UrlValue.query.filter_by(
        subject_type=stype, subject_id=subject_id, url_column_id=column_id,
    ).first()
    if uv:
        uv.url = url
    else:
        db.session.add(UrlValue(
            subject_type=stype, subject_id=subject_id,
            url_column_id=column_id, url=url,
        ))
    db.session.commit()
    return jsonify(ok=True, value=url or "")


@modules_bp.route("/modules/<int:module_id>/save-comment", methods=["POST"])
@login_required
def save_comment(module_id):
    module = get_module_or_403(module_id)
    data = request.get_json(silent=True) or {}
    subject_id = data.get("subject_id")
    comment = str(data.get("value", "")).strip() or None

    if not _subject_is_active(module, subject_id):
        return jsonify(error="Étudiant neutralisé : saisie impossible"), 403

    if module.is_group_mode:
        group = db.session.get(Group, subject_id)
        if not group or group.module_id != module.id:
            return jsonify(error="Groupe invalide"), 400
        group.comment = comment
    else:
        enr = Enrollment.query.filter_by(
            student_id=subject_id, class_id=module.class_id,
        ).first()
        if not enr:
            return jsonify(error="Inscription introuvable"), 400
        enr.general_comment = comment
    db.session.commit()
    return jsonify(ok=True)


# --------------------------------------------------------------------------- #
# Export / import Excel des notes manuelles (colonnes jaunes) + commentaire
# --------------------------------------------------------------------------- #
def _safe_sheet_title(name):
    """Titre d'onglet valide pour Excel (max 31 car., pas de []:*?/\\)."""
    title = re.sub(r"[\[\]:\*\?/\\]", " ", name or "Notes").strip()
    return (title or "Notes")[:31]


def _export_subjects(module):
    """Sujets exportables : étudiants actifs (individuel) ou groupes."""
    return [s for s in module_subjects(module) if s.get("active", True)]


def _set_comment(module, subject_id, comment):
    if module.is_group_mode:
        group = db.session.get(Group, subject_id)
        if group and group.module_id == module.id:
            group.comment = comment
    else:
        enr = Enrollment.query.filter_by(
            student_id=subject_id, class_id=module.class_id,
        ).first()
        if enr:
            enr.general_comment = comment


@modules_bp.route("/modules/<int:module_id>/export.xlsx")
@login_required
def export_module(module_id):
    """Exporte un module en .xlsx : nom du module, nom de l'étudiant/groupe,
    ses notes manuelles (colonnes jaunes) et son commentaire."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    module = get_module_or_403(module_id)
    subjects = _export_subjects(module)
    note_cols = module.note_columns
    subject_type = _subject_type(module)

    note_map = {}
    if note_cols:
        for n in NoteValue.query.filter(
            NoteValue.subject_type == subject_type,
            NoteValue.note_column_id.in_([c.id for c in note_cols])).all():
            note_map[(n.subject_id, n.note_column_id)] = n.score
    comment_by_id = {s["id"]: s.get("comment") for s in subjects}

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(module.name)

    yellow = PatternFill("solid", fgColor="FEF9C3")
    grey = PatternFill("solid", fgColor="F1F5F9")
    bold = Font(bold=True)

    # Ligne 1 : nom du module.
    ws["B1"] = f"Module : {module.name}"
    ws["B1"].font = Font(bold=True, size=13)

    # Ligne 2 : en-têtes.
    label_head = "Groupe" if module.is_group_mode else "Étudiant"
    headers = ["ID", label_head] + [c.title for c in note_cols] + ["Commentaire"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font = bold
        if 3 <= col_idx <= 2 + len(note_cols):   # colonnes de note = jaune
            cell.fill = yellow
        elif col_idx <= 2:
            cell.fill = grey

    # Données.
    for r, subj in enumerate(subjects, start=3):
        ws.cell(row=r, column=1, value=subj["id"])
        ws.cell(row=r, column=2, value=subj["label"]).font = bold
        for j, nc in enumerate(note_cols, start=3):
            v = note_map.get((subj["id"], nc.id))
            c = ws.cell(row=r, column=j, value=v)
            c.fill = yellow
        ws.cell(row=r, column=3 + len(note_cols),
                value=comment_by_id.get(subj["id"]))

    # Mise en forme : colonne ID masquée (ne pas éditer), largeurs, gel.
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 26
    for j in range(3, 3 + len(note_cols)):
        ws.column_dimensions[ws.cell(row=2, column=j).column_letter].width = 14
    ws.column_dimensions[ws.cell(row=2, column=3 + len(note_cols)).column_letter].width = 40
    ws.freeze_panes = "C3"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = re.sub(r"[^\w\-]+", "_", module.name).strip("_") or "module"
    return send_file(
        bio, as_attachment=True, download_name=f"notes_{filename}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@modules_bp.route("/modules/<int:module_id>/import", methods=["POST"])
@login_required
def import_module(module_id):
    """Réimporte un fichier .xlsx exporté : met à jour les notes manuelles et
    les commentaires. Le matching se fait par la colonne ID ; les colonnes de
    note sont reconnues par leur intitulé."""
    from openpyxl import load_workbook

    module = get_module_or_403(module_id)
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Veuillez fournir un fichier .xlsx (exporté depuis ce module).", "error")
        return redirect(url_for("modules.view_module", module_id=module.id))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        flash("Fichier Excel illisible.", "error")
        return redirect(url_for("modules.view_module", module_id=module.id))
    ws = wb.active

    rows = list(ws.iter_rows())
    # Repère la ligne d'en-tête (celle contenant "ID").
    header_row_idx = None
    for i, row in enumerate(rows):
        values = [(c.value if c.value is not None else "") for c in row]
        if any(str(v).strip().lower() == "id" for v in values):
            header_row_idx = i
            headers = [str(v).strip() for v in values]
            break
    if header_row_idx is None:
        flash("En-tête introuvable : utilisez un fichier exporté depuis ce module.", "error")
        return redirect(url_for("modules.view_module", module_id=module.id))

    # Index des colonnes.
    id_col = next(i for i, h in enumerate(headers) if h.lower() == "id")
    comment_col = next((i for i, h in enumerate(headers) if h.lower() == "commentaire"), None)
    note_by_title = {c.title: c for c in module.note_columns}
    note_cols_pos = {i: note_by_title[h] for i, h in enumerate(headers) if h in note_by_title}

    subject_type = _subject_type(module)
    valid_ids = {s["id"] for s in _export_subjects(module)}
    existing = {
        (nv.subject_id, nv.note_column_id): nv
        for nv in NoteValue.query.filter(
            NoteValue.subject_type == subject_type,
            NoteValue.note_column_id.in_([c.id for c in module.note_columns] or [0])).all()
    }

    n_notes, n_comments, errors = 0, 0, []
    for row in rows[header_row_idx + 1:]:
        cells = list(row)
        raw_id = cells[id_col].value if id_col < len(cells) else None
        if raw_id in (None, ""):
            continue
        try:
            sid = int(raw_id)
        except (TypeError, ValueError):
            continue
        if sid not in valid_ids:
            continue

        # Notes manuelles.
        for pos, nc in note_cols_pos.items():
            if pos >= len(cells):
                continue
            val = cells[pos].value
            score = None
            if val not in (None, ""):
                try:
                    score = float(str(val).replace(",", "."))
                except ValueError:
                    errors.append(f"« {nc.title} » ligne ID {sid} : valeur non numérique")
                    continue
                if not (0 <= score <= 20):
                    errors.append(f"« {nc.title} » ligne ID {sid} : hors 0–20")
                    continue
            nv = existing.get((sid, nc.id))
            if nv:
                nv.score = score
            else:
                db.session.add(NoteValue(
                    subject_type=subject_type, subject_id=sid,
                    note_column_id=nc.id, score=score,
                ))
            n_notes += 1

        # Commentaire.
        if comment_col is not None and comment_col < len(cells):
            cval = cells[comment_col].value
            _set_comment(module, sid, (str(cval).strip() if cval not in (None, "") else None))
            n_comments += 1

    db.session.commit()
    msg = f"Import terminé : {n_notes} note(s) et {n_comments} commentaire(s) mis à jour."
    if errors:
        msg += " Ignoré(s) : " + " ; ".join(errors[:5])
        if len(errors) > 5:
            msg += f" … (+{len(errors) - 5})"
        flash(msg, "error")
    else:
        flash(msg, "success")
    return redirect(url_for("modules.view_module", module_id=module.id))
