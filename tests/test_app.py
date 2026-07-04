"""Tests de bout en bout des flux clés d'AstraNote."""
from io import BytesIO

from openpyxl import load_workbook

from astranote import create_app, grading
from astranote.models import (
    db, School, AcademicYear, Class, Module, Student, Enrollment,
    GradeDate, StarColumn, NoteColumn, Group, Star, NoteValue,
)
from conftest import make_teacher, login, ADMIN_PW, TestConfig


# --------------------------------------------------------------------------- #
# Helpers de construction
# --------------------------------------------------------------------------- #
def bootstrap_class(app, admin, students=("Alice", "Bob", "Chloe"), work_mode="individual"):
    admin.post("/schools", data={"name": "EPSI"})
    admin.post("/years", data={"label": "2025-2026"})
    with app.app_context():
        sid = School.query.first().id
        yid = AcademicYear.query.first().id
    admin.post("/classes/new", data={"name": "B3", "school_id": sid,
                                     "academic_year_id": yid, "teacher_id": 1})
    with app.app_context():
        cid = Class.query.first().id
    for n in students:
        admin.post(f"/classes/{cid}/students", data={"full_name": n})
    admin.post(f"/classes/{cid}/modules/new", data={"name": "Crypto", "work_mode": work_mode})
    with app.app_context():
        mid = Module.query.first().id
        ids = {s.full_name: s.id for s in Student.query.all()}
        enr = {e.student.full_name: e.id for e in Class.query.get(cid).enrollments}
    return cid, mid, ids, enr


def add_star_column(app, admin, mid):
    admin.post(f"/modules/{mid}/dates", data={"date": "2025-09-30"})
    with app.app_context():
        did = GradeDate.query.first().id
    admin.post(f"/dates/{did}/star-columns", data={"title": "A"})
    with app.app_context():
        return did, StarColumn.query.order_by(StarColumn.id.desc()).first().id


# --------------------------------------------------------------------------- #
# Auth & CSRF
# --------------------------------------------------------------------------- #
def test_login_required(client):
    assert client.get("/", follow_redirects=False).status_code == 302


def test_csrf_field_rendered(client):
    assert b"csrf_token" in client.get("/login").data


# --------------------------------------------------------------------------- #
# Notation : prorata R1–R10
# --------------------------------------------------------------------------- #
def test_prorata_and_rounding(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    _, scid = add_star_column(app, admin, mid)
    admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Alice"], "column_id": scid, "value": "4"})
    admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Bob"], "column_id": scid, "value": "2"})
    r = admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Chloe"], "column_id": scid, "value": "3"})
    g = r.get_json()["grades"]
    assert g[str(ids["Alice"])]["note"] == 20.0 and g[str(ids["Alice"])]["is_reference"]
    assert g[str(ids["Bob"])]["note"] == 10.0
    assert g[str(ids["Chloe"])]["note"] == 15.0


def test_round_half():
    assert grading.round_half(3 / 7 * 20) == 8.5
    assert grading.round_half(2 / 3 * 20) == 13.5


def test_na_when_no_stars(app, admin):
    _, mid, _, _ = bootstrap_class(app, admin)
    add_star_column(app, admin, mid)
    assert b"N/A" in admin.get(f"/modules/{mid}").data


# --------------------------------------------------------------------------- #
# Neutralisation
# --------------------------------------------------------------------------- #
def test_neutralized_excluded_and_locked(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    _, scid = add_star_column(app, admin, mid)
    for n, v in [("Alice", "4"), ("Bob", "2"), ("Chloe", "3")]:
        admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids[n], "column_id": scid, "value": v})
    admin.post(f"/enrollments/{enr['Alice']}/toggle-active")
    # Saisie bloquée pour Alice (neutralisée)
    r = admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Alice"], "column_id": scid, "value": "1"})
    assert r.status_code == 403
    # Recalcul : Chloe devient référence, Bob 2/3*20 = 13.5
    r = admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Bob"], "column_id": scid, "value": "2"})
    g = r.get_json()["grades"]
    assert g[str(ids["Alice"])]["note"] == "—"
    assert g[str(ids["Chloe"])]["note"] == 20.0
    assert g[str(ids["Bob"])]["note"] == 13.5


# --------------------------------------------------------------------------- #
# Périmètre écoles/années
# --------------------------------------------------------------------------- #
def test_school_scoping(app, admin):
    make_teacher(app, "Prof A", "a@x.fr")
    admin.post("/schools", data={"name": "Commune"})  # admin => commune (NULL)
    ca = app.test_client()
    login(ca, "a@x.fr")
    ca.post("/schools", data={"name": "EcoleA"})
    html = ca.get("/schools").get_data(as_text=True)
    assert "EcoleA" in html and "Commune" in html
    # Un 2e prof ne voit pas EcoleA
    make_teacher(app, "Prof B", "b@x.fr")
    cb = app.test_client()
    login(cb, "b@x.fr")
    assert "EcoleA" not in cb.get("/schools").get_data(as_text=True)


# --------------------------------------------------------------------------- #
# Nettoyage des orphelins
# --------------------------------------------------------------------------- #
def test_orphan_cleanup_on_student_delete(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    _, scid = add_star_column(app, admin, mid)
    admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Alice"], "column_id": scid, "value": "3"})
    with app.app_context():
        assert Star.query.filter_by(subject_type="student", subject_id=ids["Alice"]).count() == 1
    admin.post(f"/enrollments/{enr['Alice']}/delete")
    with app.app_context():
        assert Star.query.filter_by(subject_type="student", subject_id=ids["Alice"]).count() == 0
        assert db.session.get(Student, ids["Alice"]) is None


def test_orphan_cleanup_on_group_delete(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin, work_mode="group")
    admin.post(f"/modules/{mid}/groups", data={"name": "G1"})
    with app.app_context():
        gid = Group.query.first().id
    _, scid = add_star_column(app, admin, mid)
    admin.post(f"/modules/{mid}/save-star", json={"subject_id": gid, "column_id": scid, "value": "2"})
    with app.app_context():
        assert Star.query.filter_by(subject_type="group", subject_id=gid).count() == 1
    admin.post(f"/groups/{gid}/delete")
    with app.app_context():
        assert Star.query.filter_by(subject_type="group", subject_id=gid).count() == 0


# --------------------------------------------------------------------------- #
# Renommage / réorganisation
# --------------------------------------------------------------------------- #
def test_rename_and_reorder_columns(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    admin.post(f"/modules/{mid}/dates", data={"date": "2025-09-30"})
    with app.app_context():
        did = GradeDate.query.first().id
    admin.post(f"/dates/{did}/star-columns", data={"title": "A"})
    admin.post(f"/dates/{did}/star-columns", data={"title": "B"})
    with app.app_context():
        cols = StarColumn.query.order_by(StarColumn.position).all()
        a_id, b_id = cols[0].id, cols[1].id
    admin.post(f"/star-columns/{a_id}/rename", data={"title": "Alpha"})
    admin.post(f"/star-columns/{a_id}/move", data={"dir": "down"})
    with app.app_context():
        cols = StarColumn.query.order_by(StarColumn.position).all()
        assert cols[0].id == b_id and cols[1].id == a_id
        assert db.session.get(StarColumn, a_id).title == "Alpha"


# --------------------------------------------------------------------------- #
# Édition module
# --------------------------------------------------------------------------- #
def test_edit_module_discord(app, admin):
    _, mid, _, _ = bootstrap_class(app, admin)
    admin.post(f"/modules/{mid}/edit", data={"name": "Crypto+", "discord_url": "https://discord.com/z"})
    with app.app_context():
        m = db.session.get(Module, mid)
        assert m.name == "Crypto+" and m.discord_url == "https://discord.com/z"


# --------------------------------------------------------------------------- #
# Export / import Excel
# --------------------------------------------------------------------------- #
def test_excel_round_trip(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    admin.post(f"/modules/{mid}/note-columns", data={"title": "Note CC"})
    with app.app_context():
        ncid = NoteColumn.query.first().id
    r = admin.get(f"/modules/{mid}/export.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    # trouve la ligne d'Alice et écrit sa note
    for row in range(3, 8):
        if ws.cell(row=row, column=1).value == ids["Alice"]:
            ws.cell(row=row, column=3).value = 14.5
            ws.cell(row=row, column=4).value = "Bien"
            break
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    admin.post(f"/modules/{mid}/import", data={"file": (out, "n.xlsx")},
               content_type="multipart/form-data")
    with app.app_context():
        nv = NoteValue.query.filter_by(subject_type="student", subject_id=ids["Alice"],
                                       note_column_id=ncid).first()
        assert nv.score == 14.5
        comment = Enrollment.query.filter_by(student_id=ids["Alice"], class_id=cid).first().general_comment
        assert comment == "Bien"


def test_import_rejects_out_of_range(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    admin.post(f"/modules/{mid}/note-columns", data={"title": "Note CC"})
    with app.app_context():
        ncid = NoteColumn.query.first().id
    r = admin.get(f"/modules/{mid}/export.xlsx")
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    for row in range(3, 8):
        if ws.cell(row=row, column=1).value == ids["Bob"]:
            ws.cell(row=row, column=3).value = 25   # hors 0-20
            break
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = admin.post(f"/modules/{mid}/import", data={"file": (out, "n.xlsx")},
                      content_type="multipart/form-data", follow_redirects=True)
    assert "hors 0" in resp.get_data(as_text=True)
    with app.app_context():
        assert NoteValue.query.filter_by(subject_id=ids["Bob"], note_column_id=ncid).first() is None


# --------------------------------------------------------------------------- #
# Tableau de bord : avancement de saisie
# --------------------------------------------------------------------------- #
def test_notes_sent_tracking(app, admin):
    _, mid, _, _ = bootstrap_class(app, admin)
    # Marque comme envoyées par mail à une date donnée.
    admin.post(f"/modules/{mid}/notes-sent", data={
        "notes_sent": "on", "notes_sent_date": "2026-07-01",
        "notes_sent_method": "mail", "notes_sent_detail": "au secrétariat",
    })
    with app.app_context():
        m = db.session.get(Module, mid)
        assert m.notes_sent is True
        assert m.notes_sent_date.isoformat() == "2026-07-01"
        assert m.notes_sent_method == "mail"
        assert m.notes_sent_detail == "au secrétariat"
    assert "Notes envoyées" in admin.get(f"/modules/{mid}").get_data(as_text=True)
    # Décoche : tout est réinitialisé.
    admin.post(f"/modules/{mid}/notes-sent", data={})
    with app.app_context():
        m = db.session.get(Module, mid)
        assert m.notes_sent is False and m.notes_sent_date is None
        assert m.notes_sent_method is None
    # Moyen invalide ignoré.
    admin.post(f"/modules/{mid}/notes-sent", data={"notes_sent": "on", "notes_sent_method": "pigeon"})
    with app.app_context():
        assert db.session.get(Module, mid).notes_sent_method is None


def test_dashboard_shows_notes_sent(app, admin):
    _, mid, _, _ = bootstrap_class(app, admin)
    html = admin.get("/").get_data(as_text=True)
    assert "check-unsent" in html          # module listé, non envoyé
    admin.post(f"/modules/{mid}/notes-sent", data={"notes_sent": "on", "notes_sent_date": "2026-07-01"})
    html = admin.get("/").get_data(as_text=True)
    assert "check-sent" in html and "✔" in html


def test_notes_sent_defaults_to_today(app, admin):
    from datetime import date
    _, mid, _, _ = bootstrap_class(app, admin)
    admin.post(f"/modules/{mid}/notes-sent", data={"notes_sent": "on"})  # sans date
    with app.app_context():
        assert db.session.get(Module, mid).notes_sent_date == date.today()


def test_change_password(app, admin):
    # Mauvais mot de passe actuel -> refusé
    r = admin.post("/account", data={"current_password": "wrong",
                                     "new_password": "newpass12", "confirm_password": "newpass12"},
                   follow_redirects=True)
    assert "incorrect" in r.get_data(as_text=True)
    # Trop court -> refusé
    r = admin.post("/account", data={"current_password": ADMIN_PW,
                                     "new_password": "court", "confirm_password": "court"},
                   follow_redirects=True)
    assert "8 caractères" in r.get_data(as_text=True)
    # Correct -> le nouveau mot de passe fonctionne
    admin.post("/account", data={"current_password": ADMIN_PW,
                                 "new_password": "newpass12", "confirm_password": "newpass12"})
    c2 = app.test_client()
    r = c2.post("/login", data={"email": "admin@astranote.local", "password": "newpass12"},
                follow_redirects=True)
    assert "Tableau de bord" in r.get_data(as_text=True)


def test_rename_school_and_year(app, admin):
    admin.post("/schools", data={"name": "EPSI"})
    admin.post("/years", data={"label": "2025-2026"})
    with app.app_context():
        sid = School.query.first().id
        yid = AcademicYear.query.first().id
    admin.post(f"/schools/{sid}/rename", data={"name": "EPSI Lille"})
    admin.post(f"/years/{yid}/rename", data={"label": "2026-2027"})
    with app.app_context():
        assert db.session.get(School, sid).name == "EPSI Lille"
        assert db.session.get(AcademicYear, yid).label == "2026-2027"


def test_teacher_cannot_rename_common_school(app, admin):
    admin.post("/schools", data={"name": "Commune"})  # admin => commune (NULL)
    with app.app_context():
        sid = School.query.first().id
    make_teacher(app, "Prof", "p@x.fr")
    c = app.test_client()
    login(c, "p@x.fr")
    assert c.post(f"/schools/{sid}/rename", data={"name": "Pirate"}).status_code == 403


def test_school_billing_and_class_rate(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    with app.app_context():
        sid = School.query.first().id
    # Infos de facturation de l'école
    admin.post(f"/schools/{sid}/details", data={
        "billing_emails": "compta@epsi.fr, direction@epsi.fr",
        "observation": "Payer sous 30 jours.",
    })
    # Taux horaire de la classe (virgule décimale acceptée)
    admin.post(f"/classes/{cid}/billing", data={"hourly_rate": "55,5"})
    with app.app_context():
        s = db.session.get(School, sid)
        assert s.billing_emails == "compta@epsi.fr, direction@epsi.fr"
        assert s.observation == "Payer sous 30 jours."
        assert db.session.get(Class, cid).hourly_rate == 55.5
    # Affichage sur la fiche de classe
    html = admin.get(f"/classes/{cid}").get_data(as_text=True)
    assert "compta@epsi.fr" in html and "55" in html


def test_invalid_hourly_rate_rejected(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    admin.post(f"/classes/{cid}/billing", data={"hourly_rate": "55"})
    r = admin.post(f"/classes/{cid}/billing", data={"hourly_rate": "abc"}, follow_redirects=True)
    assert "invalide" in r.get_data(as_text=True)
    with app.app_context():
        assert db.session.get(Class, cid).hourly_rate == 55  # inchangé


def test_teacher_cannot_edit_common_school_billing(app, admin):
    admin.post("/schools", data={"name": "Commune"})
    with app.app_context():
        sid = School.query.first().id
    make_teacher(app, "Prof", "p@x.fr")
    c = app.test_client()
    login(c, "p@x.fr")
    assert c.post(f"/schools/{sid}/details", data={"billing_emails": "x@x.fr"}).status_code == 403


def test_secure_cookie_flags(app):
    assert app.config["SESSION_COOKIE_HTTPONLY"] is True
    assert app.config["SESSION_COOKIE_SAMESITE"] == "Lax"


# --------------------------------------------------------------------------- #
# Administration : sauvegarde de la base
# --------------------------------------------------------------------------- #
def test_admin_page_access_control(app, admin):
    assert admin.get("/admin").status_code == 200
    make_teacher(app, "Prof", "p@x.fr")
    c = app.test_client()
    login(c, "p@x.fr")
    assert c.get("/admin").status_code == 403
    assert c.get("/admin/download-db").status_code == 403


def test_download_db_returns_sqlite_file(tmp_path):
    class FileConfig(TestConfig):
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{tmp_path / 'astranote.db'}"

    a = create_app(FileConfig)
    c = a.test_client()
    c.post("/login", data={"email": "admin@astranote.local", "password": ADMIN_PW})
    r = c.get("/admin/download-db")
    assert r.status_code == 200
    # Un fichier SQLite valide commence par cet en-tête.
    assert r.data[:16] == b"SQLite format 3\x00"
    assert "attachment" in r.headers.get("Content-Disposition", "")


def test_dashboard_progress(app, admin):
    cid, mid, ids, enr = bootstrap_class(app, admin)
    _, scid = add_star_column(app, admin, mid)
    # 1 colonne × 3 étudiants = 3 cellules ; on en remplit 1 => ~33%
    admin.post(f"/modules/{mid}/save-star", json={"subject_id": ids["Alice"], "column_id": scid, "value": "3"})
    with app.app_context():
        from astranote.main import class_saisie_progress
        klass = db.session.get(Class, cid)
        assert class_saisie_progress(klass) == 33
